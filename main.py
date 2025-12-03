"""
VendorFlow Backend API
FastAPI server with all automation logic from z-Ultimate-FIXED-FINAL.py
Self-contained - no external dependencies on other Python files
"""

import os
import json
import time
import datetime
import logging
import asyncio
import re
import requests
import math
import traceback
import tempfile
import shutil
from typing import Optional, List, Dict, Any
from pathlib import Path

from fastapi import FastAPI, HTTPException, UploadFile, File, Form, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel
import pandas as pd
from openpyxl import Workbook, load_workbook

# Selenium imports
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException

from PIL import Image
from io import BytesIO
from fake_useragent import UserAgent

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================

logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger('VendorFlow')

# In-memory log buffer for SSE streaming
log_buffer: List[Dict[str, Any]] = []
MAX_LOG_BUFFER = 1000

def add_log(message: str, level: str = "info"):
    """Add log entry to buffer and logger"""
    entry = {
        "id": len(log_buffer),
        "timestamp": datetime.datetime.now().strftime("%H:%M:%S"),
        "message": message,
        "level": level
    }
    log_buffer.append(entry)
    if len(log_buffer) > MAX_LOG_BUFFER:
        log_buffer.pop(0)
    
    if level == "error":
        logger.error(message)
    elif level == "warning":
        logger.warning(message)
    else:
        logger.info(message)

# =============================================================================
# CONFIGURATION
# =============================================================================

CONFIG_FILE = "vendor_automation_config.json"

DEFAULT_CONFIG = {
    "paths": {
        "process_folder": "/home/mkpie/vendor_data/process",
        "final_folder": "/home/mkpie/vendor_data/final",
    },
    "sellercloud": {
        "api_url": "https://an.api.sellercloud.com",
        "rest_endpoint": "https://an.api.sellercloud.com/rest",
        "username": "",
        "password": "",
        "rate_limit_per_hour": 10800,
        "delay_between_uploads": 120
    },
    "shopify": {
        "channel_name": "Shopify",
        "store_url": "",
        "api_token": ""
    },
    "eniture": {
        "api_key": "",
        "api_url": "https://s-web-api.eniture.com"
    },
    "automation": {
        "auto_scrape": True,
        "auto_process": False,
        "auto_upload": False,
        "auto_publish": False,
        "update_existing_files": True
    },
    "scraping": {
        "variation_mode": "None",
        "model_column": "Mfr Model",
        "prefix": "",
        "start_row": 1,
        "end_row": 1000,
        "save_interval": 5
    }
}

class Config:
    """Configuration management"""
    
    def __init__(self):
        self.config = self.load_config()
    
    def load_config(self) -> dict:
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r') as f:
                    loaded = json.load(f)
                    # Merge with defaults to ensure all keys exist
                    merged = DEFAULT_CONFIG.copy()
                    for key in loaded:
                        if isinstance(loaded[key], dict) and key in merged:
                            merged[key].update(loaded[key])
                        else:
                            merged[key] = loaded[key]
                    return merged
            except Exception as e:
                logger.error(f"Error loading config: {e}")
                return DEFAULT_CONFIG.copy()
        return DEFAULT_CONFIG.copy()
    
    def save_config(self) -> bool:
        try:
            with open(CONFIG_FILE, 'w') as f:
                json.dump(self.config, f, indent=4)
            return True
        except Exception as e:
            logger.error(f"Error saving config: {e}")
            return False
    
    def get(self, key_path: str, default=None):
        keys = key_path.split('.')
        value = self.config
        for key in keys:
            if isinstance(value, dict) and key in value:
                value = value[key]
            else:
                return default
        return value
    
    def set(self, key_path: str, value):
        keys = key_path.split('.')
        config = self.config
        for key in keys[:-1]:
            if key not in config:
                config[key] = {}
            config = config[key]
        config[keys[-1]] = value

# Global config instance
config = Config()

# =============================================================================
# PROCESSING STATE
# =============================================================================

class ProcessingState:
    """Track processing state across all operations"""
    def __init__(self):
        self.is_processing = False
        self.should_stop = False
        self.current_task = ""
        self.progress = 0
        self.total = 0
        self.results = []
    
    def start(self, task: str, total: int = 0):
        self.is_processing = True
        self.should_stop = False
        self.current_task = task
        self.progress = 0
        self.total = total
        self.results = []
        add_log(f"Starting {task}...", "info")
    
    def stop(self):
        self.should_stop = True
        add_log(f"Stop requested for {self.current_task}", "warning")
    
    def complete(self):
        self.is_processing = False
        add_log(f"Completed {self.current_task}", "success")
    
    def update(self, progress: int, message: str = ""):
        self.progress = progress
        if message:
            add_log(message, "info")

state = ProcessingState()

# =============================================================================
# KATOM SCRAPER
# =============================================================================

class KatomScraper:
    """Web scraper for Katom.com"""
    
    MODEL_VARIATIONS = ['LP', 'NG', '115', '1151', '120', '1201', '208', '2081', '2083', 
                        '220', '2201', '2203', '230', '2301', '2303', '240', '2401', '2403', 
                        '4403', '4003']
    GAS_VARIATIONS = ['LP', 'NG']
    ELECTRIC_VARIATIONS = ['115', '1151', '120', '1201', '208', '2081', '2083', 
                           '220', '2201', '2203', '230', '2301', '2303', '240', '2401', '2403', 
                           '4403', '4003']
    
    def __init__(self):
        self.driver = None
        self.should_stop = False
        self.session_timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def setup_driver(self) -> bool:
        """Setup Chrome driver with headless mode"""
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument(f'user-agent={UserAgent().random}')
        
        try:
            self.driver = webdriver.Chrome(options=options)
            self.driver.set_page_load_timeout(30)
            return True
        except Exception as e:
            add_log(f"Error setting up driver: {e}", "error")
            return False
    
    def close_driver(self):
        """Close WebDriver safely"""
        if self.driver:
            try:
                self.driver.quit()
            except Exception as e:
                add_log(f"Error closing driver: {e}", "error")
            self.driver = None
    
    def stop(self):
        """Stop scraping"""
        self.should_stop = True
        self.close_driver()
    
    def check_image_size(self, image_url: str) -> bool:
        """Check if image is larger than 300x300"""
        try:
            response = requests.get(image_url, timeout=10)
            if response.status_code != 200:
                return False
            img = Image.open(BytesIO(response.content))
            width, height = img.size
            return width >= 300 and height >= 300
        except Exception:
            return False
    
    def extract_numeric_price(self, price_text: str) -> str:
        """Extract numeric price from text"""
        if not price_text:
            return ""
        price_match = re.search(r'[\d,]+\.\d{2}', price_text)
        if price_match:
            return price_match.group(0).replace(',', '')
        price_match = re.search(r'[\d,]+', price_text)
        if price_match:
            return price_match.group(0).replace(',', '')
        return ""
    
    def clean_katom_links(self, html_content: str) -> str:
        """Clean katom.com links from HTML"""
        if not html_content:
            return html_content
        
        # Replace Prop 65 link
        prop65_pattern = r'<a\s+href="https://www\.katom\.com/learning-center/katom-california-proposition-65-advisories\.html"[^>]*>([^<]*)</a>'
        prop65_replacement = '<a href="https://www.cityfoodequipment.com/pages/city-food-equipment-california-proposition-65.html" title="City Food Equipment Prop 65 Info">Prop 65 information</a>'
        html_content = re.sub(prop65_pattern, prop65_replacement, html_content, flags=re.IGNORECASE)
        
        # Remove free shipping text
        free_shipping_pattern = r'\*[^*]+is free shipping eligible\. Users must be signed in with an order total greater than \$500 and be shipping to the 48 contiguous states\.'
        html_content = re.sub(free_shipping_pattern, '', html_content, flags=re.IGNORECASE)
        
        # Remove katom.com links
        katom_link_pattern = r'<a\s+[^>]*href="[^"]*katom\.com[^"]*"[^>]*>.*?</a>'
        html_content = re.sub(katom_link_pattern, '', html_content, flags=re.IGNORECASE | re.DOTALL)
        
        return html_content
    
    def process_weight_value(self, value) -> str:
        """Process weight values - add 5 to rounded value"""
        try:
            number_match = re.search(r'(\d+(\.\d+)?)', str(value))
            if number_match:
                number = float(number_match.group(1))
                rounded = math.ceil(number)
                final = rounded + 5
                units_match = re.search(r'[^\d.]+$', str(value))
                units = units_match.group(0).strip() if units_match else ""
                return f"{final}{' ' + units if units else ''}"
            return value
        except:
            return value
    
    def extract_table_data(self) -> tuple:
        """Extract specs table data"""
        specs_dict = {}
        specs_html = ""
        try:
            specs_tables = self.driver.find_elements(By.CSS_SELECTOR, "table.table.table-condensed.specs-table")
            if not specs_tables:
                specs_tables = self.driver.find_elements(By.TAG_NAME, "table")
            if specs_tables:
                table = specs_tables[0]
                rows = table.find_elements(By.TAG_NAME, "tr")
                specs_html = '<div style="overflow-x: auto;"><table class="specs-table" cellspacing="0" cellpadding="4" border="1" style="margin-top:10px;border-collapse:collapse;width:auto;" align="left"><tbody>'
                for row in rows:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if len(cells) >= 2:
                        key = cells[0].text.strip()
                        cell_html = cells[1].get_attribute('innerHTML')
                        value_text = cells[1].text.strip()
                        
                        cell_html = re.sub(r'<img[^>]*>', '', cell_html)
                        cell_html = self.clean_katom_links(cell_html)
                        
                        if "weight" in key.lower():
                            value = self.process_weight_value(value_text)
                        else:
                            value = value_text
                        
                        if key and key.lower() not in specs_dict:
                            specs_dict[key.lower()] = value
                        
                        specs_html += f'<tr><td style="padding:3px 8px;"><b>{key}</b></td><td style="padding:3px 8px;">{cell_html}</td></tr>'
                
                specs_html += "</tbody></table></div>"
        except Exception as e:
            add_log(f"Error extracting table data: {e}", "error")
        return specs_dict, specs_html
    
    def extract_carousel_images(self) -> List[str]:
        """Extract images from carousel"""
        carousel_images = []
        try:
            carousel_container = self.driver.find_element(By.CSS_SELECTOR, "div.product-image-container div.carousel-inner")
            carousel_items = carousel_container.find_elements(By.CSS_SELECTOR, "div.carousel-item")
            for item in carousel_items:
                images = item.find_elements(By.TAG_NAME, "img")
                for img in images:
                    src = img.get_attribute("src")
                    if src and (src.startswith("http://") or src.startswith("https://") or src.startswith("//")):
                        if src not in carousel_images:
                            carousel_images.append(src)
        except NoSuchElementException:
            pass
        except Exception as e:
            add_log(f"Error extracting carousel images: {e}", "error")
        return carousel_images
    
    def scrape_katom(self, model_number: str, prefix: str, variant_suffix: str = "", retries: int = 2) -> tuple:
        """Scrape Katom for a specific model"""
        if self.should_stop:
            return False, None
        
        model_number = ''.join(e for e in model_number if e.isalnum()).upper()
        if model_number.endswith("HC"):
            model_number = model_number[:-2]
        
        full_model = f"{model_number}{variant_suffix}" if variant_suffix else model_number
        url = f"https://www.katom.com/{prefix}-{full_model}.html"
        
        if not self.setup_driver():
            return False, None
        
        title, description = "Title not found", ""
        specs_dict, specs_html = {}, ""
        video_links, numeric_price, main_image = "", "", ""
        additional_images = []
        item_found = False
        
        try:
            self.driver.get(url)
            
            if "404" in self.driver.title or "not found" in self.driver.title.lower():
                return False, None
            
            # Extract title
            try:
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "h1.product-name.mb-0, h1"))
                )
                title_element = self.driver.find_element(By.CSS_SELECTOR, "h1.product-name.mb-0")
                title = title_element.text.strip()
                if title:
                    item_found = True
            except TimeoutException:
                try:
                    for selector in ["h1.product-title", "h1[itemprop='name']", "h1"]:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                        if elements:
                            title = elements[0].text.strip()
                            item_found = True
                            break
                except Exception:
                    pass
            except Exception:
                pass
            
            if not item_found:
                return False, None
            
            # Extract price
            try:
                price_elements = self.driver.find_elements(By.CSS_SELECTOR, "p.product-price-text.m-0")
                if price_elements:
                    for element in price_elements:
                        price_text = element.text.strip()
                        if price_text:
                            numeric_price = self.extract_numeric_price(price_text)
                            if numeric_price:
                                break
                
                if not numeric_price:
                    for selector in [".product-price", ".price", "[class*='price']"]:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                        if elements:
                            for element in elements:
                                price_text = element.text.strip()
                                if price_text and ('$' in price_text or re.search(r'\d+\.\d{2}', price_text)):
                                    numeric_price = self.extract_numeric_price(price_text)
                                    if numeric_price:
                                        break
                            if numeric_price:
                                break
            except Exception:
                pass
            
            # Extract carousel images
            carousel_images = self.extract_carousel_images()
            if carousel_images:
                main_image = carousel_images[0]
                additional_images.extend(carousel_images[1:])
            
            # Extract additional images
            try:
                for selector in [".additional-images img", ".product-thumbnails img", ".thumb-image"]:
                    elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                    if elements:
                        for element in elements:
                            src = element.get_attribute("src")
                            full_size_src = src
                            if src and "thumbnail" in src.lower():
                                full_size_src = src.replace("thumbnail", "full")
                            if full_size_src and full_size_src != main_image and full_size_src not in additional_images:
                                if self.check_image_size(full_size_src):
                                    additional_images.append(full_size_src)
                                    if len(additional_images) >= 5:
                                        break
                        if len(additional_images) >= 5:
                            break
            except Exception:
                pass
            
            # Fallback main image
            if not main_image:
                try:
                    for selector in [".product-img img", ".main-product-image", "img.main-image"]:
                        elements = self.driver.find_elements(By.CSS_SELECTOR, selector)
                        if elements:
                            for element in elements:
                                src = element.get_attribute("src")
                                if src and self.check_image_size(src):
                                    main_image = src
                                    break
                            if main_image:
                                break
                except Exception:
                    pass
            
            # Extract description
            try:
                custom_content = self.driver.find_elements(By.CLASS_NAME, "product-custom-content")
                filtered = []
                
                if custom_content:
                    paragraphs = custom_content[0].find_elements(By.TAG_NAME, "p")
                    for p in paragraphs:
                        p_html = p.get_attribute('innerHTML').strip()
                        p_text = p.text.strip()
                        if p_text and not p_text.lower().startswith("*free") and "video" not in p_text.lower():
                            p_html = re.sub(r'<img[^>]*>', '', p_html)
                            p_html = self.clean_katom_links(p_html)
                            filtered.append(f"<p>{p_html}</p>")
                else:
                    tab_content = self.driver.find_element(By.CLASS_NAME, "tab-content")
                    paragraphs = tab_content.find_elements(By.TAG_NAME, "p")
                    for p in paragraphs:
                        p_html = p.get_attribute('innerHTML').strip()
                        p_text = p.text.strip()
                        if p_text and not p_text.lower().startswith("*free") and "video" not in p_text.lower():
                            p_html = re.sub(r'<img[^>]*>', '', p_html)
                            p_html = self.clean_katom_links(p_html)
                            filtered.append(f"<p>{p_html}</p>")
                
                description = "".join(filtered) if filtered else ""
            except Exception:
                pass
            
            # Extract specs
            specs_dict, specs_html = self.extract_table_data()
            
            # Extract video links
            try:
                sources = self.driver.find_elements(By.CSS_SELECTOR, "source[src*='.mp4'], source[type*='video']")
                for source in sources:
                    src = source.get_attribute("src")
                    if src and src not in video_links:
                        video_links += f"{src}\n"
            except Exception:
                pass
            
        except Exception as e:
            add_log(f"Error scraping {full_model}: {e}", "error")
            if retries > 0:
                self.close_driver()
                time.sleep(2)
                return self.scrape_katom(model_number, prefix, variant_suffix, retries - 1)
            return False, None
        finally:
            self.close_driver()
        
        result_data = {
            'title': title,
            'description': description,
            'specs_dict': specs_dict,
            'specs_html': specs_html,
            'video_links': video_links,
            'numeric_price': numeric_price,
            'main_image': main_image,
            'additional_images': additional_images
        }
        
        return True, result_data
    
    def determine_auto_variation_mode(self, model: str, spec_mapping: dict) -> str:
        """Determine variation mode based on AQ Specification"""
        if model not in spec_mapping:
            return "None"
        
        spec = spec_mapping[model]
        has_gas = bool(re.search(r'\bgas\b', spec, re.IGNORECASE))
        has_electric = bool(re.search(r'\belectric\b', spec, re.IGNORECASE))
        
        if has_gas and has_electric:
            return "Check All"
        elif has_gas:
            return "Gas Type"
        elif has_electric:
            return "Electric"
        return "None"
    
    def scrape_with_variations(self, model_number: str, prefix: str, variation_mode: str) -> List[tuple]:
        """Scrape model with variations"""
        results = []
        
        # Try original
        add_log(f"Trying original model: {model_number}")
        success, data = self.scrape_katom(model_number, prefix, variant_suffix="")
        if success and data:
            results.append(("", data))
            add_log(f"Found original: {model_number}", "success")
        
        # Determine variations
        if variation_mode == "None":
            variations_to_try = []
        elif variation_mode == "Gas Type":
            variations_to_try = self.GAS_VARIATIONS
        elif variation_mode == "Electric":
            variations_to_try = self.ELECTRIC_VARIATIONS
        elif variation_mode == "Low Voltage":
            variations_to_try = ["115", "1151", "120", "1201", "208", "2081", "220", "2201", "230", "2301", "240", "2401"]
        elif variation_mode == "Check All":
            variations_to_try = self.MODEL_VARIATIONS
        else:
            variations_to_try = []
        
        for variant in variations_to_try:
            if self.should_stop:
                break
            add_log(f"Trying variation: {model_number}{variant}")
            success, data = self.scrape_katom(model_number, prefix, variant_suffix=variant)
            if success and data:
                results.append((variant, data))
                add_log(f"Found variation: {model_number}{variant}", "success")
            time.sleep(0.5)
        
        return results
    
    def scrape_models(self, models: List[str], variation_mode: str, prefix: str, 
                      start_row: int = 1, end_row: int = 1000, 
                      csv_path: str = None) -> List[dict]:
        """Scrape multiple models"""
        self.should_stop = False
        
        # Build spec mapping for Auto mode
        spec_mapping = {}
        if variation_mode == "Auto" and csv_path and os.path.exists(csv_path):
            try:
                df = pd.read_csv(csv_path)
                if "AQ Specification" in df.columns:
                    model_col = None
                    for col in df.columns:
                        if "model" in col.lower() and "mfr" in col.lower():
                            model_col = col
                            break
                    if model_col:
                        for idx, row in df.iterrows():
                            model_val = str(row.get(model_col, "")).strip()
                            spec_val = str(row.get("AQ Specification", "")).strip().lower()
                            if model_val:
                                spec_mapping[model_val] = spec_val
            except Exception as e:
                add_log(f"Error loading CSV for Auto mode: {e}", "error")
        
        all_rows = []
        end_row = min(end_row, len(models))
        total_models = end_row - start_row + 1
        
        for i in range(start_row - 1, end_row):
            if self.should_stop:
                break
            
            model = models[i]
            
            # Determine actual variation mode
            actual_mode = variation_mode
            if variation_mode == "Auto":
                actual_mode = self.determine_auto_variation_mode(model, spec_mapping)
                add_log(f"Auto mode for {model}: Using '{actual_mode}'")
            
            variation_results = self.scrape_with_variations(model, prefix, actual_mode)
            
            if not variation_results:
                add_log(f"No results for model: {model}", "warning")
                state.update(i - start_row + 2, f"No results for {model}")
                continue
            
            for variant_suffix, data in variation_results:
                price_value = data['numeric_price'] if data['numeric_price'] else "Call for Price"
                
                combined_description = f'<div style="text-align: justify;">{data["description"]}</div>'
                if data['specs_html']:
                    combined_description += f'<h3 style="margin-top: 15px;">Specifications</h3>{data["specs_html"]}'
                
                new_row = {
                    "Mfr Model": model,
                    "Model Variant": variant_suffix if variant_suffix else "Original",
                    "Title": data['title'],
                    "Description": combined_description,
                    "Price": price_value,
                    "Main Image": data['main_image']
                }
                
                for idx, img_url in enumerate(data['additional_images'][:5], 1):
                    new_row[f"Additional Image {idx}"] = img_url
                
                video_list = [link.strip() for link in data['video_links'].strip().split('\n') if link.strip()]
                for idx, link in enumerate(video_list[:5], 1):
                    new_row[f"Video Link {idx}"] = link
                
                for key, value in data['specs_dict'].items():
                    col_name = key.title()
                    new_row[col_name] = value
                
                all_rows.append(new_row)
            
            progress = int(((i - start_row + 1) / total_models) * 100)
            state.update(progress, f"Scraped {model}: {len(variation_results)} results")
        
        return all_rows


# =============================================================================
# TAG PROCESSOR
# =============================================================================

class TagProcessor:
    """Tag processor for Shopify"""
    
    MODEL_VARIATIONS = ['LP', 'NG', '1151', '1201', '2081', '2083', '2201', '2203', 
                        '2301', '2303', '2401', '2403', '4403', '4003']
    SKIP_HEADERS = ['Additional Image 1', 'Additional Image 2', 'Additional Image 3', 
                    'Additional Image 4', 'Additional Image 5', 'Video Link 1']
    
    def __init__(self):
        self.should_stop = False
        self.shopify_store = config.get('shopify.store_url')
        self.shopify_token = config.get('shopify.api_token')
    
    def stop(self):
        self.should_stop = True
    
    def clean_sku(self, sku: str) -> str:
        """Remove variant suffixes from SKU"""
        if not sku:
            return sku
        cleaned = sku
        pattern = r'-(LP|NG|1151|1201|2081|2083|2201|2203|2301|2303|2401|2403|4403|4003)$'
        while re.search(pattern, cleaned):
            cleaned = re.sub(pattern, '', cleaned)
        return cleaned
    
    def get_product_from_shopify(self, sku: str) -> Optional[dict]:
        """Search Shopify for product by SKU using GraphQL"""
        try:
            url = f"https://{self.shopify_store}/admin/api/2024-01/graphql.json"
            headers = {
                'X-Shopify-Access-Token': self.shopify_token,
                'Content-Type': 'application/json'
            }
            
            clean_sku = str(sku).strip()
            query = """
            query getProductBySku($sku: String!) {
                productVariants(first: 1, query: $sku) {
                    edges {
                        node {
                            id
                            sku
                            product { id }
                        }
                    }
                }
            }
            """
            
            payload = {'query': query, 'variables': {'sku': clean_sku}}
            time.sleep(0.5)
            
            response = requests.post(url, headers=headers, json=payload, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                if 'errors' in data:
                    return None
                
                edges = data.get('data', {}).get('productVariants', {}).get('edges', [])
                if edges:
                    node = edges[0]['node']
                    variant_sku = node.get('sku', '')
                    
                    if variant_sku.upper() == clean_sku.upper():
                        variant_gid = node.get('id', '')
                        product_gid = node.get('product', {}).get('id', '')
                        
                        variant_id = variant_gid.split('/')[-1] if variant_gid else None
                        product_id = product_gid.split('/')[-1] if product_gid else None
                        
                        if product_id and variant_id:
                            return {
                                'productId': str(product_id),
                                'variantId': str(variant_id),
                                'productGid': product_gid
                            }
            return None
        except Exception as e:
            add_log(f"Error getting product from Shopify: {e}", "error")
            return None
    
    def update_product_tags(self, product_gid: str, tags: str) -> bool:
        """Update product tags in Shopify"""
        try:
            url = f"https://{self.shopify_store}/admin/api/2024-01/graphql.json"
            headers = {
                'X-Shopify-Access-Token': self.shopify_token,
                'Content-Type': 'application/json'
            }
            
            tag_list = [tag.strip() for tag in tags.split(',') if tag.strip()]
            
            mutation = """
            mutation updateProductTags($input: ProductInput!) {
                productUpdate(input: $input) {
                    product { id tags }
                    userErrors { field message }
                }
            }
            """
            
            payload = {
                'query': mutation,
                'variables': {'input': {'id': product_gid, 'tags': tag_list}}
            }
            
            time.sleep(0.5)
            response = requests.post(url, headers=headers, json=payload, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                if 'errors' in data:
                    return False
                user_errors = data.get('data', {}).get('productUpdate', {}).get('userErrors', [])
                return len(user_errors) == 0
            return False
        except Exception as e:
            add_log(f"Error updating tags: {e}", "error")
            return False
    
    def process_tags(self, excel_path: str, csv_path: str, output_path: str) -> dict:
        """Process tags from Excel and save to CSV"""
        try:
            add_log("Loading Excel file...")
            excel_workbook = load_workbook(excel_path, data_only=True)
            worksheet = excel_workbook.active
            
            # Build lookup table
            add_log("Building lookup table...")
            column_headers = {}
            for col in range(12, 98):
                header_cell = worksheet.cell(row=1, column=col)
                if header_cell.value:
                    header_text = str(header_cell.value).strip()
                    if header_text:
                        column_headers[col] = header_text
            
            lookup_table = {}
            max_row = worksheet.max_row
            
            for row_idx in range(2, max_row + 1):
                if self.should_stop:
                    raise Exception("Stopped by user")
                
                mfr_model_cell = worksheet.cell(row=row_idx, column=1)
                mfr_model = str(mfr_model_cell.value).strip() if mfr_model_cell.value else ""
                
                if mfr_model and mfr_model != "None":
                    tag_values = []
                    for col in range(12, 98):
                        cell_value = worksheet.cell(row=row_idx, column=col)
                        cell_text = str(cell_value.value).strip() if cell_value.value else ""
                        if cell_text and cell_text != "None" and cell_text != "NaN":
                            if col in column_headers:
                                header_name = column_headers[col]
                                if header_name not in self.SKIP_HEADERS:
                                    tag_values.append(f"{header_name}: {cell_text}")
                            else:
                                tag_values.append(cell_text)
                    lookup_table[mfr_model] = ", ".join(tag_values) if tag_values else ""
                
                if row_idx % 100 == 0:
                    state.update(int((row_idx / max_row) * 40), f"Processing Excel row {row_idx}/{max_row}")
            
            excel_workbook.close()
            add_log(f"Lookup table: {len(lookup_table)} entries")
            
            # Load and update CSV
            add_log("Loading Shopify CSV...")
            csv_data = pd.read_csv(csv_path, encoding='utf-8')
            
            update_count = 0
            total_rows = len(csv_data)
            
            for idx, row in csv_data.iterrows():
                if self.should_stop:
                    raise Exception("Stopped by user")
                
                variant_sku = str(row.get('Variant SKU', '')).strip() if pd.notna(row.get('Variant SKU')) else ""
                if variant_sku and variant_sku != "None":
                    cleaned_sku = self.clean_sku(variant_sku)
                    if cleaned_sku in lookup_table:
                        csv_data.at[idx, 'Tags'] = lookup_table[cleaned_sku]
                        update_count += 1
                
                if (idx + 1) % 100 == 0:
                    state.update(40 + int((idx / total_rows) * 55), f"Updated {update_count} rows")
            
            # Export
            add_log(f"Exporting to {output_path}...")
            csv_data.to_csv(output_path, index=False, encoding='utf-8')
            
            return {
                'success': True,
                'updated_rows': update_count,
                'total_rows': total_rows,
                'output_path': output_path,
                'lookup_entries': len(lookup_table)
            }
        except Exception as e:
            add_log(f"Error in tag processing: {e}", "error")
            return {'success': False, 'error': str(e)}
    
    def push_tags_to_shopify(self, file_path: str) -> dict:
        """Push tags directly to Shopify API"""
        try:
            add_log("Loading file...")
            if file_path.lower().endswith(('.xlsx', '.xls')):
                file_data = pd.read_excel(file_path)
            else:
                file_data = pd.read_csv(file_path, encoding='utf-8')
            
            all_columns = list(file_data.columns)
            tag_columns = all_columns[7:] if len(all_columns) > 7 else []
            skip_headers = ['Additional Image 1', 'Additional Image 2', 'Additional Image 3', 
                           'Additional Image 4', 'Additional Image 5', 'Video Link 1']
            tag_columns = [col for col in tag_columns if col not in skip_headers]
            
            successful_updates = 0
            failed_updates = 0
            skipped_rows = 0
            total_rows = len(file_data)
            errors = []
            
            for idx, row in file_data.iterrows():
                if self.should_stop:
                    raise Exception("Stopped by user")
                
                mfr_model = str(row.get('Mfr Model', '')).strip() if pd.notna(row.get('Mfr Model')) else ""
                model_variant = str(row.get('Model Variant', '')).strip() if pd.notna(row.get('Model Variant')) else ""
                
                if not mfr_model:
                    skipped_rows += 1
                    continue
                
                if model_variant and model_variant.lower() != "original":
                    sku = f"{mfr_model}-{model_variant}"
                else:
                    sku = mfr_model
                
                tag_values = []
                for col in tag_columns:
                    cell_value = str(row.get(col, '')).strip() if pd.notna(row.get(col)) else ""
                    if cell_value and cell_value != "None" and cell_value != "NaN":
                        tag_values.append(f"{col}: {cell_value}")
                
                if not tag_values:
                    skipped_rows += 1
                    continue
                
                tags = ", ".join(tag_values)
                shopify_product = self.get_product_from_shopify(sku)
                
                if not shopify_product:
                    failed_updates += 1
                    errors.append(f"SKU not found: {sku}")
                    continue
                
                if self.update_product_tags(shopify_product['productGid'], tags):
                    successful_updates += 1
                    add_log(f"Updated tags for {sku}", "success")
                else:
                    failed_updates += 1
                    errors.append(f"Failed to update: {sku}")
                
                if (idx + 1) % 10 == 0:
                    state.update(int((idx / total_rows) * 100), 
                               f"Processing {idx + 1}/{total_rows}")
            
            return {
                'success': True,
                'successful_updates': successful_updates,
                'failed_updates': failed_updates,
                'skipped_rows': skipped_rows,
                'total_rows': total_rows,
                'errors': errors[:20]
            }
        except Exception as e:
            add_log(f"Error pushing tags: {e}", "error")
            return {'success': False, 'error': str(e)}


# =============================================================================
# WEIGHT PROCESSOR
# =============================================================================

class WeightProcessor:
    """Weight and dimensions processor"""
    
    MODEL_VARIATIONS = ['LP', 'NG', '1151', '1201', '2081', '2083', '2201', '2203', 
                        '2301', '2303', '2401', '2403', '4403', '4003']
    
    def __init__(self):
        self.should_stop = False
    
    def stop(self):
        self.should_stop = True
    
    def clean_sku(self, sku: str) -> str:
        if not sku:
            return sku
        cleaned = sku
        pattern = r'-(LP|NG|1151|1201|2081|2083|2201|2203|2301|2303|2401|2403|4403|4003)$'
        while re.search(pattern, cleaned):
            cleaned = re.sub(pattern, '', cleaned)
        return cleaned
    
    def process_weights(self, vendor_csv_path: str, output_csv_path: str) -> dict:
        """Process weights from vendor CSV"""
        try:
            add_log("Loading vendor CSV...")
            vendor_data_df = pd.read_csv(vendor_csv_path, encoding='utf-8')
            
            # Find columns
            vendor_cols = vendor_data_df.columns.tolist()
            col_map = {}
            for col in vendor_cols:
                col_lower = col.strip().lower()
                if col_lower == "mfr model":
                    col_map['model'] = col
                elif col_lower == "shipping weight":
                    col_map['weight'] = col
                elif col_lower == "width":
                    col_map['width'] = col
                elif col_lower == "depth":
                    col_map['depth'] = col
                elif col_lower == "height":
                    col_map['height'] = col
                elif col_lower == "freight class":
                    col_map['freight_class'] = col
                elif col_lower == "ship from zip":
                    col_map['ship_zip'] = col
            
            # Build lookup
            add_log("Building vendor lookup...")
            vendor_data = {}
            for idx, row in vendor_data_df.iterrows():
                if self.should_stop:
                    raise Exception("Stopped by user")
                
                mfr_model = str(row.get(col_map.get('model', ''), '')).strip()
                if mfr_model and mfr_model != "None":
                    try:
                        weight = float(row.get(col_map.get('weight', ''), 0) or 0)
                    except:
                        weight = 0
                    try:
                        width = float(row.get(col_map.get('width', ''), 0) or 0)
                    except:
                        width = 0
                    try:
                        depth = float(row.get(col_map.get('depth', ''), 0) or 0)
                    except:
                        depth = 0
                    try:
                        height = float(row.get(col_map.get('height', ''), 0) or 0)
                    except:
                        height = 0
                    
                    vendor_data[mfr_model] = {
                        'weight': weight,
                        'width': width,
                        'depth': depth,
                        'height': height,
                        'freight_class': row.get(col_map.get('freight_class', ''), ''),
                        'ship_zip': row.get(col_map.get('ship_zip', ''), '')
                    }
                
                if (idx + 1) % 100 == 0:
                    state.update(int((idx / len(vendor_data_df)) * 40), f"Building lookup: {idx + 1}")
            
            add_log(f"Vendor lookup: {len(vendor_data)} entries")
            
            # Load output CSV
            add_log("Loading output CSV...")
            csv_data = pd.read_csv(output_csv_path, encoding='utf-8')
            
            update_count = 0
            total_rows = len(csv_data)
            
            for idx, row in csv_data.iterrows():
                if self.should_stop:
                    raise Exception("Stopped by user")
                
                variant_sku = str(row.get('Variant SKU', '')).strip() if pd.notna(row.get('Variant SKU')) else ""
                if variant_sku and variant_sku != "None":
                    cleaned_sku = self.clean_sku(variant_sku)
                    if cleaned_sku in vendor_data:
                        info = vendor_data[cleaned_sku]
                        
                        weight_val = info['weight']
                        csv_data.at[idx, 'Weight'] = math.ceil(weight_val) if weight_val > 0 else weight_val
                        csv_data.at[idx, 'Length'] = math.ceil(info['width']) if info['width'] > 0 else info['width']
                        csv_data.at[idx, 'Width'] = math.ceil(info['depth']) if info['depth'] > 0 else info['depth']
                        csv_data.at[idx, 'Height'] = math.ceil(info['height']) if info['height'] > 0 else info['height']
                        csv_data.at[idx, 'Freight Class'] = info['freight_class']
                        csv_data.at[idx, 'Dropship Zipcode'] = info['ship_zip']
                        
                        # Quote Method
                        if pd.notna(weight_val) and weight_val > 0:
                            csv_data.at[idx, 'Quote Method'] = 'S' if weight_val < 85 else 'L'
                        else:
                            csv_data.at[idx, 'Quote Method'] = 'S' if pd.isna(info['freight_class']) or info['freight_class'] == '' else 'L'
                        
                        # Default freight class
                        if (pd.isna(csv_data.at[idx, 'Freight Class']) or csv_data.at[idx, 'Freight Class'] == '') and csv_data.at[idx, 'Quote Method'] == 'L':
                            csv_data.at[idx, 'Freight Class'] = 175
                        
                        update_count += 1
                
                if (idx + 1) % 100 == 0:
                    state.update(40 + int((idx / total_rows) * 55), f"Updated {update_count} rows")
            
            add_log(f"Exporting to {output_csv_path}...")
            csv_data.to_csv(output_csv_path, index=False, encoding='utf-8')
            
            return {
                'success': True,
                'updated_rows': update_count,
                'total_rows': total_rows,
                'output_path': output_csv_path,
                'vendor_entries': len(vendor_data)
            }
        except Exception as e:
            add_log(f"Error processing weights: {e}", "error")
            return {'success': False, 'error': str(e)}


# =============================================================================
# ENITURE SYNC
# =============================================================================

class EnitureSync:
    """Sync to Eniture API"""
    
    def __init__(self):
        self.should_stop = False
        self.shopify_store = config.get('shopify.store_url')
        self.shopify_token = config.get('shopify.api_token')
        self.eniture_key = config.get('eniture.api_key')
        self.eniture_url = config.get('eniture.api_url')
    
    def stop(self):
        self.should_stop = True
    
    def get_variant_from_shopify(self, sku: str) -> Optional[dict]:
        """Get variant from Shopify by SKU"""
        try:
            url = f"https://{self.shopify_store}/admin/api/2024-01/graphql.json"
            headers = {
                'X-Shopify-Access-Token': self.shopify_token,
                'Content-Type': 'application/json'
            }
            
            query = """
            query getProductBySku($sku: String!) {
                productVariants(first: 1, query: $sku) {
                    edges {
                        node {
                            id
                            sku
                            product { id }
                        }
                    }
                }
            }
            """
            
            payload = {'query': query, 'variables': {'sku': str(sku).strip()}}
            time.sleep(0.5)
            
            response = requests.post(url, headers=headers, json=payload, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                edges = data.get('data', {}).get('productVariants', {}).get('edges', [])
                if edges:
                    node = edges[0]['node']
                    if node.get('sku', '').upper() == sku.upper():
                        return {
                            'productId': node['product']['id'].split('/')[-1],
                            'variantId': node['id'].split('/')[-1]
                        }
            return None
        except Exception as e:
            add_log(f"Shopify lookup error: {e}", "error")
            return None
    
    def sync_to_eniture(self, lookup_path: str, weight_csv_path: str) -> dict:
        """Sync to Eniture API"""
        try:
            add_log("Loading lookup file...")
            if lookup_path.endswith('.xlsx'):
                lookup_df = pd.read_excel(lookup_path)
            else:
                lookup_df = pd.read_csv(lookup_path, encoding='utf-8')
            
            add_log("Loading weight CSV...")
            weight_csv = pd.read_csv(weight_csv_path, encoding='utf-8')
            
            # Find columns
            product_id_col = None
            manufacturer_sku_col = None
            for col in lookup_df.columns:
                col_lower = col.strip().lower()
                if col_lower == "productid":
                    product_id_col = col
                elif col_lower == "manufacturersku":
                    manufacturer_sku_col = col
            
            if not product_id_col or not manufacturer_sku_col:
                raise Exception("Lookup file must have ProductID and ManufacturerSKU columns")
            
            mfr_model_col = None
            for col in weight_csv.columns:
                if col.strip().lower() == "mfr model":
                    mfr_model_col = col
                    break
            
            if not mfr_model_col:
                raise Exception("Weight CSV must have 'Mfr Model' column")
            
            total_rows = len(lookup_df)
            successful = 0
            failed = 0
            skipped = 0
            errors = []
            
            eniture_headers = {
                'Authorization': f'Bearer {self.eniture_key}',
                'Accept': 'application/json',
                'X-Shopify-Shop': self.shopify_store,
                'Content-Type': 'application/json'
            }
            
            for idx, lookup_row in lookup_df.iterrows():
                if self.should_stop:
                    raise Exception("Stopped by user")
                
                product_id = str(lookup_row.get(product_id_col, '')).strip()
                manufacturer_sku = str(lookup_row.get(manufacturer_sku_col, '')).strip()
                
                if not product_id or not manufacturer_sku or product_id == "None":
                    skipped += 1
                    continue
                
                # Find weight data
                weight_row = None
                for _, wrow in weight_csv.iterrows():
                    w_model = str(wrow.get(mfr_model_col, '')).strip()
                    if w_model == manufacturer_sku:
                        weight_row = wrow
                        break
                
                if weight_row is None:
                    failed += 1
                    errors.append(f"No weight data for: {manufacturer_sku}")
                    continue
                
                # Get Shopify variant
                shopify_variant = self.get_variant_from_shopify(product_id)
                if not shopify_variant:
                    failed += 1
                    errors.append(f"Shopify lookup failed: {product_id}")
                    continue
                
                # Prepare payload
                try:
                    weight = float(weight_row.get('Shipping Weight', 0) or 0)
                except:
                    weight = 0
                try:
                    length = float(weight_row.get('Depth', 0) or 0)
                except:
                    length = 0
                try:
                    width = float(weight_row.get('Width', 0) or 0)
                except:
                    width = 0
                try:
                    height = float(weight_row.get('Height', 0) or 0)
                except:
                    height = 0
                try:
                    freight_class = float(weight_row.get('Freight Class', 0) or 0)
                except:
                    freight_class = 0
                
                # Quote method
                quote_method = 'S' if weight > 0 and weight < 85 else 'L'
                if weight == 0:
                    quote_method = 'S' if freight_class == 0 else 'L'
                
                # Freight class validation
                valid_classes = [50, 55, 60, 65, 70, 77.5, 85, 92.5, 100, 110, 125, 150, 175, 200, 250, 300, 400, 500]
                if freight_class in valid_classes:
                    freight_class_str = str(int(freight_class)) if freight_class not in [77.5, 92.5] else str(freight_class)
                else:
                    freight_class_str = "DensityBased"
                
                if (freight_class == 0) and quote_method == 'L':
                    freight_class_str = "175"
                
                payload = {
                    "data": {
                        "productId": shopify_variant['productId'],
                        "variantId": shopify_variant['variantId'],
                        "attributes": {
                            "quoteMethod": quote_method,
                            "weight": max(1, math.ceil(weight)),
                            "freightClass": freight_class_str,
                            "width": max(1, math.ceil(width)),
                            "height": max(1, math.ceil(height)),
                            "length": max(1, math.ceil(length))
                        }
                    }
                }
                
                try:
                    response = requests.post(
                        f"{self.eniture_url}/api/products",
                        headers=eniture_headers,
                        json=payload,
                        timeout=10
                    )
                    
                    if response.status_code == 200:
                        successful += 1
                        add_log(f"Synced: {product_id}", "success")
                    else:
                        failed += 1
                        errors.append(f"API error for {product_id}: {response.status_code}")
                except Exception as e:
                    failed += 1
                    errors.append(f"Request failed for {product_id}: {str(e)}")
                
                if (idx + 1) % 10 == 0:
                    state.update(int((idx / total_rows) * 100), 
                               f"Synced: {successful}, Failed: {failed}")
            
            return {
                'success': True,
                'successful_syncs': successful,
                'failed_syncs': failed,
                'skipped_rows': skipped,
                'total_rows': total_rows,
                'errors': errors[:20]
            }
        except Exception as e:
            add_log(f"Eniture sync error: {e}", "error")
            return {'success': False, 'error': str(e)}


# =============================================================================
# FASTAPI APPLICATION
# =============================================================================

app = FastAPI(title="VendorFlow API", version="1.0.0")

# CORS for frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Ensure upload directories exist
UPLOAD_DIR = Path("/tmp/vendorflow_uploads")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

# =============================================================================
# PYDANTIC MODELS
# =============================================================================

class ScrapingRequest(BaseModel):
    models: List[str]
    variation_mode: str = "None"
    prefix: str
    start_row: int = 1
    end_row: int = 1000

class ConfigUpdate(BaseModel):
    config: dict

class TagProcessRequest(BaseModel):
    excel_path: str
    csv_path: str
    output_path: str

class WeightProcessRequest(BaseModel):
    vendor_path: str
    output_path: str

class EnitureSyncRequest(BaseModel):
    lookup_path: str
    weight_csv_path: str

# =============================================================================
# API ENDPOINTS
# =============================================================================

@app.get("/")
async def root():
    return {"status": "ok", "service": "VendorFlow API", "version": "1.0.0"}

@app.get("/api/status")
async def get_status():
    return {
        "is_processing": state.is_processing,
        "current_task": state.current_task,
        "progress": state.progress,
        "total": state.total
    }

@app.post("/api/stop")
async def stop_processing():
    state.stop()
    return {"status": "stop_requested"}

# --- CONFIG ---

@app.get("/api/config")
async def get_config():
    return config.config

@app.post("/api/config")
async def update_config(update: ConfigUpdate):
    for key, value in update.config.items():
        if isinstance(value, dict):
            for subkey, subvalue in value.items():
                config.set(f"{key}.{subkey}", subvalue)
        else:
            config.set(key, value)
    
    if config.save_config():
        return {"status": "saved"}
    raise HTTPException(status_code=500, detail="Failed to save config")

# --- FILE UPLOAD ---

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    try:
        file_path = UPLOAD_DIR / file.filename
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        return {"path": str(file_path), "filename": file.filename}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- LOGS SSE ---

@app.get("/api/logs/stream")
async def stream_logs():
    async def event_generator():
        last_id = 0
        while True:
            new_logs = [log for log in log_buffer if log['id'] >= last_id]
            for log in new_logs:
                yield f"data: {json.dumps(log)}\n\n"
                last_id = log['id'] + 1
            await asyncio.sleep(0.5)
    
    return StreamingResponse(event_generator(), media_type="text/event-stream")

@app.get("/api/logs")
async def get_logs(limit: int = 100):
    return log_buffer[-limit:]

# --- SCRAPING ---

@app.post("/api/scrape")
async def start_scraping(request: ScrapingRequest, background_tasks: BackgroundTasks):
    if state.is_processing:
        raise HTTPException(status_code=409, detail="Already processing")
    
    def run_scraping():
        state.start("Web Scraping", len(request.models))
        scraper = KatomScraper()
        try:
            results = scraper.scrape_models(
                request.models,
                request.variation_mode,
                request.prefix,
                request.start_row,
                request.end_row
            )
            state.results = results
            state.complete()
        except Exception as e:
            add_log(f"Scraping error: {e}", "error")
            state.is_processing = False
        finally:
            scraper.close_driver()
    
    background_tasks.add_task(run_scraping)
    return {"status": "started", "task": "scraping"}

@app.post("/api/scrape/file")
async def scrape_from_file(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    model_column: str = Form("Mfr Model"),
    prefix: str = Form(...),
    variation_mode: str = Form("None"),
    start_row: int = Form(1),
    end_row: int = Form(1000)
):
    if state.is_processing:
        raise HTTPException(status_code=409, detail="Already processing")
    
    # Save uploaded file
    file_path = UPLOAD_DIR / file.filename
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Read models from CSV
    try:
        df = pd.read_csv(file_path)
        if model_column not in df.columns:
            raise HTTPException(status_code=400, detail=f"Column '{model_column}' not found")
        models = df[model_column].dropna().astype(str).tolist()
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    def run_scraping():
        state.start("Web Scraping", len(models))
        scraper = KatomScraper()
        try:
            results = scraper.scrape_models(
                models, variation_mode, prefix, start_row, 
                min(end_row, len(models)), str(file_path)
            )
            state.results = results
            
            # Save results
            if results:
                output_path = UPLOAD_DIR / f"scraped_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                pd.DataFrame(results).to_excel(output_path, index=False)
                add_log(f"Results saved to {output_path}", "success")
            
            state.complete()
        except Exception as e:
            add_log(f"Scraping error: {e}", "error")
            state.is_processing = False
        finally:
            scraper.close_driver()
    
    background_tasks.add_task(run_scraping)
    return {"status": "started", "models_count": len(models)}

@app.get("/api/scrape/results")
async def get_scraping_results():
    return {"results": state.results, "count": len(state.results)}

# --- TAGS ---

@app.post("/api/tags/process")
async def process_tags(
    background_tasks: BackgroundTasks,
    excel_file: UploadFile = File(...),
    csv_file: UploadFile = File(...),
    output_name: str = Form("tags_output.csv")
):
    if state.is_processing:
        raise HTTPException(status_code=409, detail="Already processing")
    
    # Save files
    excel_path = UPLOAD_DIR / excel_file.filename
    csv_path = UPLOAD_DIR / csv_file.filename
    output_path = UPLOAD_DIR / output_name
    
    with open(excel_path, "wb") as f:
        shutil.copyfileobj(excel_file.file, f)
    with open(csv_path, "wb") as f:
        shutil.copyfileobj(csv_file.file, f)
    
    def run_processing():
        state.start("Tag Processing")
        processor = TagProcessor()
        try:
            result = processor.process_tags(str(excel_path), str(csv_path), str(output_path))
            state.results = [result]
            state.complete()
        except Exception as e:
            add_log(f"Tag processing error: {e}", "error")
            state.is_processing = False
    
    background_tasks.add_task(run_processing)
    return {"status": "started", "output": str(output_path)}

@app.post("/api/tags/push")
async def push_tags_to_shopify(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...)
):
    if state.is_processing:
        raise HTTPException(status_code=409, detail="Already processing")
    
    file_path = UPLOAD_DIR / file.filename
    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    
    def run_push():
        state.start("Tag Push to Shopify")
        processor = TagProcessor()
        try:
            result = processor.push_tags_to_shopify(str(file_path))
            state.results = [result]
            state.complete()
        except Exception as e:
            add_log(f"Tag push error: {e}", "error")
            state.is_processing = False
    
    background_tasks.add_task(run_push)
    return {"status": "started"}

# --- WEIGHTS ---

@app.post("/api/weights/process")
async def process_weights(
    background_tasks: BackgroundTasks,
    vendor_file: UploadFile = File(...),
    output_file: UploadFile = File(...)
):
    if state.is_processing:
        raise HTTPException(status_code=409, detail="Already processing")
    
    vendor_path = UPLOAD_DIR / vendor_file.filename
    output_path = UPLOAD_DIR / output_file.filename
    
    with open(vendor_path, "wb") as f:
        shutil.copyfileobj(vendor_file.file, f)
    with open(output_path, "wb") as f:
        shutil.copyfileobj(output_file.file, f)
    
    def run_processing():
        state.start("Weight Processing")
        processor = WeightProcessor()
        try:
            result = processor.process_weights(str(vendor_path), str(output_path))
            state.results = [result]
            state.complete()
        except Exception as e:
            add_log(f"Weight processing error: {e}", "error")
            state.is_processing = False
    
    background_tasks.add_task(run_processing)
    return {"status": "started"}

# --- ENITURE ---

@app.post("/api/eniture/sync")
async def sync_to_eniture(
    background_tasks: BackgroundTasks,
    lookup_file: UploadFile = File(...),
    weight_file: UploadFile = File(...)
):
    if state.is_processing:
        raise HTTPException(status_code=409, detail="Already processing")
    
    lookup_path = UPLOAD_DIR / lookup_file.filename
    weight_path = UPLOAD_DIR / weight_file.filename
    
    with open(lookup_path, "wb") as f:
        shutil.copyfileobj(lookup_file.file, f)
    with open(weight_path, "wb") as f:
        shutil.copyfileobj(weight_file.file, f)
    
    def run_sync():
        state.start("Eniture Sync")
        sync = EnitureSync()
        try:
            result = sync.sync_to_eniture(str(lookup_path), str(weight_path))
            state.results = [result]
            state.complete()
        except Exception as e:
            add_log(f"Eniture sync error: {e}", "error")
            state.is_processing = False
    
    background_tasks.add_task(run_sync)
    return {"status": "started"}

# --- DOWNLOAD ---

@app.get("/api/download/{filename}")
async def download_file(filename: str):
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")
    
    return StreamingResponse(
        open(file_path, "rb"),
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    import uvicorn
    
    # Create necessary directories
    Path(config.get('paths.process_folder', '/home/mkpie/vendor_data/process')).mkdir(parents=True, exist_ok=True)
    Path(config.get('paths.final_folder', '/home/mkpie/vendor_data/final')).mkdir(parents=True, exist_ok=True)
    
    uvicorn.run(app, host="0.0.0.0", port=8000)
